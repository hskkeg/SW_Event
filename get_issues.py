from redminelib import Redmine
from src.List import *
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border

# single line comment
"""
  Multi lines comment
"""

"""
    10.19   31536 - [BR] [IDB] TM HEV NA SOP R/C event 用
    10.19   31688 - [BR] [100KR] HM 21MY M Event
    10.20   30612 - [BR] [100CN] SV51-01 GB5(MoC Si Base) SOP R/C
    10.20   31702 - [BR][IDB] RG3 SPORT P1 2차 등록 EVENT
    10.20   24215 - [BR][IDB]CN7 HEV 북미 SOP EVENT
    10.20   17010 - [BR] [100KR] NX4a M Event
    10.22   31667 - [BR] [100CN] CHANA P201 7th SOP R/C 用
    10.22   17073 - [100KR] US4 T Event
    10.23   31587 - [BR][IDB] FORD CD542 pre-R05 event
    10.23   31588 - [BR][100GM] FORD CD542 pre-R05 event
    10.23   16973 - [BR] [100KR] RJ PE P1 Event
"""
ED4_WORK_LIST = [31536, 31688, 30612, 31702, 24215, 17010, 31667, 17073, 31587, 31588, 16973]
REG_EVENT_LIST = ['TM HEV NA SOP RC', 'HM 21MY M','SV51-01 GB5 SOP RC', 'RG3 SPORT P1 2nd', 'CN7 HEV NA SOP', 'NX4a M', 'P201 7th RC', 'US4 T', 'IDB CD542 pre-R05', 'MGH CD542 pre-R05', 'RJ PE P1']

def PM_Redmine_Issue_List(Input):
    redmine = Redmine('http://191.1.11.178', username='sk.hahm', password='dbsguswls22@')
    ED4_TitleList = []
    ED4_MemberName = []
    ED4_Tracker_List = []
    ED4_URL_List = []
    s = redmine.issue.get(Input)

    for i in s.children:
        if ('Mandatory' in redmine.issue.get(i.id).tracker.name) | ('Category' in redmine.issue.get(i.id).tracker.name) | ('Carry_Over' in redmine.issue.get(i.id).tracker.name):
            i = redmine.issue.get(i.id)
            for j in i.children:
                try:
                    Name = redmine.issue.get(j.id).assigned_to.name
                    Title = j.subject
                    URL = j.url
                    Tracker = str(redmine.issue.get(i.id).tracker)

                    #if ('/' in Tracker):
                        #Tracker = Tracker.replace("/", "")

                    for Change in ED4_Find_Member():
                        if Name in Change[0]:
                            ED4_TitleList.append(Title)
                            ED4_MemberName.append(Change[-1])
                            ED4_Tracker_List.append(Tracker)
                            ED4_URL_List.append(URL)
                            print(URL)
                except:
                    pass

        else:
            try:
                Name = redmine.issue.get(i.id).assigned_to.name
                Title = i.subject
                URL = i.url
                Tracker = str(redmine.issue.get(i.id).tracker)

                #if ('/' in Tracker):
                    #Tracker = Tracker.replace("/", "")


                for Change in ED4_Find_Member():
                    if Name in Change[0]:
                        ED4_TitleList.append(Title)
                        ED4_MemberName.append(Change[-1])
                        ED4_Tracker_List.append(Tracker)
                        ED4_URL_List.append(URL)
                        print(URL)
            except:
                pass

    for idx in range(len(ED4_WORK_LIST)):
        print(idx)

        if(Input == ED4_WORK_LIST[idx]):
            WSNum = 'ws'+str(idx)
            WSNum = wb.create_sheet()
            WSNum.title = str(REG_EVENT_LIST[idx])
            Export_Excel(WSNum, ED4_TitleList, ED4_MemberName, ED4_Tracker_List, ED4_URL_List)

def ED4_Member():
    ED4member = []
    for i in ED4_MemberList():
        ED4member.append(i.split('/'))
    return ED4member

def ED4_Find_Member():
    ED4findmember = []
    for j in ED4_Member():
        ED4findmember.append(j)
    return ED4findmember

def ED4_Work_List_Func():
    for j in range(len(ED4_WORK_LIST)):
        PM_Redmine_Issue_List(ED4_WORK_LIST[j])

def Export_Excel(SheetNum, TiltleList, MemberName, Tracker, URL):
    SheetNum.cell(1, 1, '일감')
    SheetNum.cell(1, 2, '담당자')
    SheetNum.cell(1, 3, '유형')
    SheetNum.cell(1, 4, '레드마인 주소')
    for j in range(len(TiltleList)):
        SheetNum.cell(j + 2, 1, TiltleList[j])
        SheetNum.cell(j + 2, 2, MemberName[j])
        SheetNum.cell(row=j + 2, column=3).value = Tracker[j]
        SheetNum.cell(j + 2, 4, URL[j])
    #SheetNum.column_dimensions.auto_size = True
    SheetNum.column_dimensions['A'].width = 95
    SheetNum.column_dimensions['B'].width = 9
    SheetNum.column_dimensions['C'].width = 15
    SheetNum.column_dimensions['D'].width = 32


wb = Workbook()     # create work book
ED4_Work_List_Func()
wb.save("CW43.xlsx")

# PM_Redmine_Issue_List(17111)